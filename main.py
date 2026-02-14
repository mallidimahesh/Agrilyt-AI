from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from services.imd_service import get_weather_data, get_historical_weather
from services.mandi_service import get_market_prices
from services.fci_service import get_fci_stock
from services.market_intelligence_service import get_market_intelligence
from services.ml_engine import predict_future_prices
from services.blockchain_service import intervention_blockchain
import os, pathlib
from openpyxl import Workbook, load_workbook

app = FastAPI(
    title="Agrilyt Ai",
    description="Intelligent Price Forecasting & Stabilization Platform for Essential Commodities"
)

# Support CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all for hackathon demo
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Excel-based user storage ────────────────────────────────
USERS_FILE = pathlib.Path(__file__).parent / "users.xlsx"

def _ensure_workbook():
    """Create users.xlsx with headers if it doesn't exist."""
    if not USERS_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "users"
        ws.append(["name", "email", "password", "user_type"])
        wb.save(USERS_FILE)

class RegisterBody(BaseModel):
    name: str
    email: str
    password: str
    user_type: str  # consumer | farmer | govt

class LoginBody(BaseModel):
    email: str
    password: str

@app.post("/api/register")
def register(body: RegisterBody):
    _ensure_workbook()
    wb = load_workbook(USERS_FILE)
    ws = wb["users"]
    # Check for duplicate email
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[1].lower() == body.email.lower():
            raise HTTPException(status_code=400, detail="Email already registered")
    ws.append([body.name, body.email, body.password, body.user_type])
    wb.save(USERS_FILE)
    return {"message": "Registration successful", "name": body.name, "email": body.email, "user_type": body.user_type}

@app.post("/api/login")
def login(body: LoginBody):
    _ensure_workbook()
    wb = load_workbook(USERS_FILE)
    ws = wb["users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and row[1].lower() == body.email.lower() and row[2] == body.password:
            return {"message": "Login successful", "name": row[0], "email": row[1], "user_type": row[3]}
    raise HTTPException(status_code=401, detail="Invalid email or password")

@app.get("/")
def home():
    return {"message": "AgriPrice Predictor API is running"}

@app.get("/api/weather/{city_id}")
def weather(city_id: str):
    data = get_weather_data(city_id)
    return data

@app.get("/api/predict")
def predict(commodity: str, city_id: str):
    """
    Main prediction endpoint.
    Retrieves weather (history + forecast) and market history,
    then feeds into ML engine.
    """
   
    import concurrent.futures
    
    with concurrent.futures.ThreadPoolExecutor() as executor:
     
        future_weather_current = executor.submit(get_weather_data, city_id)
        future_weather_history = executor.submit(get_historical_weather, city_id)
        future_market_data = executor.submit(get_market_prices, commodity, city_id=city_id)
        future_fci = executor.submit(get_fci_stock, city_id, commodity)
        future_macro = executor.submit(get_market_intelligence, commodity)
      
        weather_current = future_weather_current.result()
        weather_history = future_weather_history.result()
        market_data = future_market_data.result()
        fci_stock = future_fci.result()
        macro_factors = future_macro.result()
    
    if not market_data or not market_data.get('records'):
        raise HTTPException(status_code=404, detail="Market data not found")

    # 2. Predict (ML Engine)
    predictions = predict_future_prices(
        weather_history=weather_history,
        weather_forecast=weather_current, 
        price_history=market_data
    )
    
    # 3. Audit (Blockchain)
    from config import Config
    if Config.ENABLE_FORECAST_AUDIT and predictions:
        try:
            from services.blockchain_service import intervention_blockchain
            from datetime import datetime
            
            # Create a summary hash of the forecast
            forecast_summary = f"{commodity}_{city_id}_{predictions[0]['predicted_price']}_{predictions[-1]['predicted_price']}"
            
            intervention_blockchain.add_new_transaction({
                "type": "FORECAST_AUDIT",
                "commodity": commodity,
                "city": city_id,
                "model": predictions[0].get('model', 'Unknown'),
                "start_price": predictions[0]['predicted_price'],
                "end_price": predictions[-1]['predicted_price'],
                "forecast_days": len(predictions),
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            # In a real app we'd mine asynchronously or periodically
            intervention_blockchain.mine()
        except Exception as b_err:
            print(f"Blockchain audit failed: {b_err}")
    
    return {
        "commodity": commodity,
        "current_weather": weather_current,
        "weather_history": weather_history,
        "market_history": market_data['records'][-30:], 
        "fci_stock": fci_stock,
        "macro_factors": macro_factors,
        "forecast": predictions
    }

@app.get("/api/blockchain/chain")
def get_chain():
    chain_data = []
    for block in intervention_blockchain.chain:
        chain_data.append(block.__dict__)
    return {"length": len(chain_data), "chain": chain_data}

@app.post("/api/blockchain/mine")
def mine_block():
    result = intervention_blockchain.mine()
    if not result:
        return {"message": "No transactions to mine"}
    return {"message": f"Block #{result} mined successfully"}

if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.getenv("PORT", 8004))
    uvicorn.run(app, host="0.0.0.0", port=port)
